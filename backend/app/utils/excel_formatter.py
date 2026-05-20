from io import BytesIO
from typing import Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
CONFIRMED_FILL = PatternFill("solid", fgColor="E2EFDA")
ISSUE_FILL = PatternFill("solid", fgColor="FFF2CC")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _fmt_dt(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return str(val)


class ExcelFormatter:

    @staticmethod
    def format_events(events: list[dict], include_client_name: bool = False) -> BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Events"

        # Collect all JSONB field keys across all events
        field_keys: list[str] = []
        seen: set[str] = set()
        for ev in events:
            for k in (ev.get("fields") or {}).keys():
                if k not in seen:
                    field_keys.append(k)
                    seen.add(k)

        headers: list[str] = []
        if include_client_name:
            headers.append("Client Name")
        headers += ["Time Summary", "Timestamp", "Source Host"]
        headers += field_keys
        headers += ["Confirmed By", "Confirmed At", "Issue",
                    "Issue Raised By", "Issue Raised At"]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"

        for row_idx, ev in enumerate(events, start=2):
            fields = ev.get("fields") or {}
            is_confirmed = bool(ev.get("confirmed_by_username"))
            has_issue = bool(ev.get("issue_text"))

            if is_confirmed:
                row_fill = CONFIRMED_FILL
            elif has_issue:
                row_fill = ISSUE_FILL
            elif row_idx % 2 == 0:
                row_fill = ALT_FILL
            else:
                row_fill = WHITE_FILL

            values: list = []
            if include_client_name:
                values.append(ev.get("client_name", ""))
            values += [
                ev.get("time_summary") or _fmt_dt(ev.get("timestamp")),
                _fmt_dt(ev.get("timestamp")),
                ev.get("source_host") or "",
            ]
            values += [fields.get(k, "") for k in field_keys]
            values += [
                ev.get("confirmed_by_username") or "",
                _fmt_dt(ev.get("confirmed_at")),
                ev.get("issue_text") or "",
                ev.get("issue_raised_by_username") or "",
                _fmt_dt(ev.get("issue_raised_at")),
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill

        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(header)
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def write_audit_log(rows: list[dict]) -> BytesIO:
        """Generate an Excel workbook for audit log rows."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Log"

        headers = [
            "ID", "Username", "Role", "Event Type", "Client ID",
            "Target ID", "Details", "IP Address", "User Agent", "Performed At",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"

        for row_idx, row in enumerate(rows, start=2):
            row_fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
            values = [
                row.get("id"),
                row.get("username") or "",
                row.get("role") or "",
                row.get("event_type") or "",
                row.get("client_id"),
                row.get("target_id"),
                row.get("details") or "",
                row.get("ip_address") or "",
                row.get("user_agent") or "",
                _fmt_dt(row.get("performed_at")),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill

        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(header)
            for row_idx in range(2, ws.max_row + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf